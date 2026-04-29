import os
import shutil
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import Department, User
from documents.exporting import build_job_exports
from documents.models import AuditLog, BatchUpload, Job, PageResult, ReviewFlag, SystemConfig
from documents.templatetags.document_filters import render_ocr_text, render_table_cell
from documents.views import ASYNC_APPROVE_ALL_THRESHOLD
from tasks.ocr_task import process_document


class HighlightRenderingTests(TestCase):
    def test_render_ocr_text_uses_span_offsets_for_duplicate_words(self):
        html = str(
            render_ocr_text(
                "TOTAL 100\nTOTAL 200",
                [
                    SimpleNamespace(
                        word="TOTAL",
                        confidence=0.42,
                        span_offset=10,
                        span_length=5,
                    )
                ],
            )
        )

        self.assertEqual(html.count("<mark"), 1)
        self.assertGreater(html.find("<mark"), html.find("100"))

    def test_render_table_cell_highlights_review_tokens(self):
        html = str(
            render_table_cell(
                {
                    "content": "Amount 2000",
                    "review_tokens": [{"word": "2000", "confidence": 0.41}],
                }
            )
        )

        self.assertIn("<mark", html)
        self.assertIn("2000", html)


class BulkUploadTests(TestCase):
    def setUp(self):
        scratch_root = os.path.join(os.path.dirname(os.path.dirname(__file__)), "_test_scratch")
        os.makedirs(scratch_root, exist_ok=True)
        self.temp_root = os.path.join(scratch_root, uuid4().hex)
        self.media_root = os.path.join(self.temp_root, "uploads")
        os.makedirs(self.media_root, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(self.temp_root, ignore_errors=True))

        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.addCleanup(self.override.disable)

        self.user = User.objects.create_user(
            username="batch-operator",
            email="batch-operator@example.com",
            password="secret123",
            role=User.Role.OPERATOR,
        )
        self.other_user = User.objects.create_user(
            username="other-user",
            email="other-user@example.com",
            password="secret123",
            role=User.Role.OPERATOR,
        )
        self.client.force_login(self.user)

    @patch("documents.views.queue_processing_job")
    def test_multi_file_upload_creates_batch_and_independent_jobs(self, queue_mock):
        response = self.client.post(
            reverse("upload"),
            {
                "document_type": Job.DocumentType.PAYROLL,
                "files": [
                    SimpleUploadedFile("batch-a.pdf", b"%PDF-1.4 fake a"),
                    SimpleUploadedFile("batch-b.pdf", b"%PDF-1.4 fake b"),
                ],
            },
        )

        self.assertEqual(queue_mock.call_count, 2)
        self.assertEqual(BatchUpload.objects.count(), 1)
        batch = BatchUpload.objects.get()
        jobs = list(Job.objects.order_by("filename"))

        self.assertEqual(len(jobs), 2)
        self.assertEqual(batch.requested_document_type, Job.DocumentType.PAYROLL)
        self.assertTrue(all(job.batch_id == batch.id for job in jobs))
        self.assertEqual({job.filename for job in jobs}, {"batch-a.pdf", "batch-b.pdf"})
        self.assertRedirects(response, reverse("batch_detail", args=[batch.id]))

    @patch("documents.views.queue_processing_job")
    def test_single_file_upload_keeps_document_detail_redirect(self, queue_mock):
        response = self.client.post(
            reverse("upload"),
            {
                "document_type": Job.DocumentType.PAYROLL,
                "files": [SimpleUploadedFile("single.pdf", b"%PDF-1.4 fake")],
            },
        )

        self.assertEqual(queue_mock.call_count, 1)
        self.assertEqual(BatchUpload.objects.count(), 0)
        job = Job.objects.get()
        self.assertIsNone(job.batch)
        self.assertRedirects(response, reverse("document_detail", args=[job.id]))

    @patch("documents.views.queue_processing_job")
    def test_invalid_multi_file_submission_rejects_entire_batch(self, queue_mock):
        response = self.client.post(
            reverse("upload"),
            {
                "document_type": Job.DocumentType.PAYROLL,
                "files": [
                    SimpleUploadedFile("valid.pdf", b"%PDF-1.4 fake"),
                    SimpleUploadedFile("invalid.txt", b"not allowed"),
                ],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "unsupported file type")
        self.assertEqual(BatchUpload.objects.count(), 0)
        self.assertEqual(Job.objects.count(), 0)
        queue_mock.assert_not_called()

    def test_batch_detail_is_scoped_to_owner(self):
        batch = BatchUpload.objects.create(
            owner=self.user,
            requested_document_type=Job.DocumentType.PAYROLL,
        )
        Job.objects.create(
            owner=self.user,
            batch=batch,
            filename="owned.pdf",
            document_type=Job.DocumentType.PAYROLL,
            status=Job.Status.PENDING,
            upload=SimpleUploadedFile("owned.pdf", b"%PDF-1.4 fake"),
        )

        owner_response = self.client.get(reverse("batch_detail", args=[batch.id]))
        self.assertEqual(owner_response.status_code, 200)

        self.client.force_login(self.other_user)
        other_response = self.client.get(reverse("batch_detail", args=[batch.id]))
        self.assertEqual(other_response.status_code, 404)


class DepartmentAccessTests(TestCase):
    def setUp(self):
        scratch_root = os.path.join(os.path.dirname(os.path.dirname(__file__)), "_test_scratch")
        os.makedirs(scratch_root, exist_ok=True)
        self.temp_root = os.path.join(scratch_root, uuid4().hex)
        self.media_root = os.path.join(self.temp_root, "uploads")
        os.makedirs(self.media_root, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(self.temp_root, ignore_errors=True))

        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.addCleanup(self.override.disable)

        self.operations = Department.objects.create(name="Operations", confidence_threshold=0.82)
        self.finance = Department.objects.create(name="Finance", confidence_threshold=0.66)

        self.supervisor = User.objects.create_user(
            username="ops-supervisor",
            email="ops-supervisor@example.com",
            password="secret123",
            role=User.Role.SUPERVISOR,
            department=self.operations,
        )
        self.operator = User.objects.create_user(
            username="ops-operator",
            email="ops-operator@example.com",
            password="secret123",
            role=User.Role.OPERATOR,
            department=self.operations,
            supervisor=self.supervisor,
        )
        self.other_operator = User.objects.create_user(
            username="finance-operator",
            email="finance-operator@example.com",
            password="secret123",
            role=User.Role.OPERATOR,
            department=self.finance,
        )
        self.admin = User.objects.create_user(
            username="admin-user",
            email="admin-user@example.com",
            password="secret123",
            role=User.Role.ADMIN,
        )

    @patch("documents.views.queue_processing_job")
    def test_operator_upload_assigns_department_to_batch_and_jobs(self, queue_mock):
        self.client.force_login(self.operator)

        response = self.client.post(
            reverse("upload"),
            {
                "document_type": Job.DocumentType.PAYROLL,
                "files": [
                    SimpleUploadedFile("department-a.pdf", b"%PDF-1.4 fake a"),
                    SimpleUploadedFile("department-b.pdf", b"%PDF-1.4 fake b"),
                ],
            },
        )

        self.assertEqual(response.status_code, 302)
        batch = BatchUpload.objects.get()
        jobs = list(Job.objects.order_by("filename"))
        self.assertEqual(batch.department, self.operations)
        self.assertTrue(all(job.department == self.operations for job in jobs))

    def test_job_effective_confidence_threshold_prefers_department(self):
        config = SystemConfig.get()
        config.confidence_threshold = 0.70
        config.save()

        job = Job.objects.create(
            owner=self.operator,
            department=self.operations,
            filename="threshold.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PENDING,
            upload=SimpleUploadedFile("threshold.pdf", b"%PDF-1.4 fake"),
        )

        self.assertEqual(job.effective_confidence_threshold, 0.82)

    def test_supervisor_document_list_is_department_scoped(self):
        own_job = Job.objects.create(
            owner=self.operator,
            department=self.operations,
            filename="operations.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.COMPLETED,
            upload=SimpleUploadedFile("operations.pdf", b"%PDF-1.4 fake"),
        )
        Job.objects.create(
            owner=self.other_operator,
            department=self.finance,
            filename="finance.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.COMPLETED,
            upload=SimpleUploadedFile("finance.pdf", b"%PDF-1.4 fake"),
        )

        self.client.force_login(self.supervisor)
        response = self.client.get(reverse("document_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own_job.filename)
        self.assertNotContains(response, "finance.pdf")

    def test_supervisor_batch_access_is_department_scoped(self):
        own_batch = BatchUpload.objects.create(
            owner=self.operator,
            department=self.operations,
            requested_document_type=Job.DocumentType.GENERIC,
        )
        Job.objects.create(
            owner=self.operator,
            department=self.operations,
            batch=own_batch,
            filename="ops-batch.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PENDING,
            upload=SimpleUploadedFile("ops-batch.pdf", b"%PDF-1.4 fake"),
        )
        other_batch = BatchUpload.objects.create(
            owner=self.other_operator,
            department=self.finance,
            requested_document_type=Job.DocumentType.GENERIC,
        )
        Job.objects.create(
            owner=self.other_operator,
            department=self.finance,
            batch=other_batch,
            filename="finance-batch.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PENDING,
            upload=SimpleUploadedFile("finance-batch.pdf", b"%PDF-1.4 fake"),
        )

        self.client.force_login(self.supervisor)
        own_response = self.client.get(reverse("batch_detail", args=[own_batch.id]))
        other_response = self.client.get(reverse("batch_detail", args=[other_batch.id]))

        self.assertEqual(own_response.status_code, 200)
        self.assertEqual(other_response.status_code, 404)

    def test_supervisor_audit_log_is_department_scoped(self):
        own_job = Job.objects.create(
            owner=self.operator,
            department=self.operations,
            filename="audited-ops.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.COMPLETED,
            upload=SimpleUploadedFile("audited-ops.pdf", b"%PDF-1.4 fake"),
        )
        other_job = Job.objects.create(
            owner=self.other_operator,
            department=self.finance,
            filename="audited-finance.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.COMPLETED,
            upload=SimpleUploadedFile("audited-finance.pdf", b"%PDF-1.4 fake"),
        )
        AuditLog.objects.create(user=self.operator, department=self.operations, job=own_job, action="upload")
        AuditLog.objects.create(user=self.other_operator, department=self.finance, job=other_job, action="upload")

        self.client.force_login(self.supervisor)
        response = self.client.get(reverse("audit_log"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.operator.username)
        self.assertContains(response, own_job.filename)
        self.assertNotContains(response, self.other_operator.username)
        self.assertNotContains(response, other_job.filename)

    def test_supervisor_can_update_department_threshold(self):
        self.client.force_login(self.supervisor)

        response = self.client.post(reverse("settings_save"), {"confidence_threshold": "90"})

        self.assertEqual(response.status_code, 302)
        self.operations.refresh_from_db()
        self.finance.refresh_from_db()
        self.assertEqual(self.operations.confidence_threshold, 0.90)
        self.assertEqual(self.finance.confidence_threshold, 0.66)

    def test_operator_cannot_access_settings_page(self):
        self.client.force_login(self.operator)

        response = self.client.get(reverse("settings"))

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("document_list"))

    def test_supervisor_cannot_stop_operator_job(self):
        job = Job.objects.create(
            owner=self.operator,
            department=self.operations,
            filename="running.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PROCESSING,
            processing_task_id="task-1",
            upload=SimpleUploadedFile("running.pdf", b"%PDF-1.4 fake"),
        )

        self.client.force_login(self.supervisor)
        response = self.client.post(reverse("stop_job", args=[job.id]))

        self.assertEqual(response.status_code, 302)
        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.PROCESSING)


class ExportRegenerationTests(TestCase):
    def setUp(self):
        scratch_root = os.path.join(os.path.dirname(os.path.dirname(__file__)), "_test_scratch")
        os.makedirs(scratch_root, exist_ok=True)
        self.temp_root = os.path.join(scratch_root, uuid4().hex)
        self.export_dir = os.path.join(self.temp_root, "exports")
        self.media_root = os.path.join(self.temp_root, "uploads")
        os.makedirs(self.export_dir, exist_ok=True)
        os.makedirs(self.media_root, exist_ok=True)
        self.addCleanup(lambda: shutil.rmtree(self.temp_root, ignore_errors=True))

        self.override = override_settings(EXPORT_DIR=self.export_dir, MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.addCleanup(self.override.disable)

        self.user = User.objects.create_user(
            username="operator",
            email="operator@example.com",
            password="secret123",
            role=User.Role.OPERATOR,
        )
        self.client.force_login(self.user)

    def test_approving_flag_regenerates_excel_without_cell_highlight(self):
        job = Job.objects.create(
            owner=self.user,
            filename="sample.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.NEEDS_REVIEW,
            upload=SimpleUploadedFile("sample.pdf", b"%PDF-1.4 fake"),
        )
        page = PageResult.objects.create(
            job=job,
            page_number=1,
            extracted_text="Amount 1000",
            tables_json=[
                [
                    [
                        {"content": "Item", "raw_content": "Item", "table_index": 1, "row_index": 0, "column_index": 0},
                        {"content": "Amount", "raw_content": "Amount", "table_index": 1, "row_index": 0, "column_index": 1},
                    ],
                    [
                        {"content": "Fee", "raw_content": "Fee", "table_index": 1, "row_index": 1, "column_index": 0},
                        {"content": "1000", "raw_content": "1000", "table_index": 1, "row_index": 1, "column_index": 1},
                    ],
                ]
            ],
        )
        flag = ReviewFlag.objects.create(
            page_result=page,
            word="1000",
            confidence=0.42,
            span_offset=7,
            span_length=4,
            table_index=1,
            row_index=1,
            column_index=1,
        )

        build_job_exports(job)
        xlsx_path = os.path.join(self.export_dir, f"{job.id}.xlsx")
        workbook = load_workbook(xlsx_path)
        self.assertIsNotNone(workbook["P1_Table_1"]["B2"].comment)
        workbook.close()

        response = self.client.post(
            reverse("approve_flags", args=[job.id]),
            {"flag_ids": str(flag.id)},
        )
        self.assertEqual(response.status_code, 302)

        workbook = load_workbook(xlsx_path)
        self.assertIsNone(workbook["P1_Table_1"]["B2"].comment)
        workbook.close()
        flag.refresh_from_db()
        job.refresh_from_db()
        self.assertTrue(flag.reviewed)
        self.assertEqual(job.status, Job.Status.COMPLETED)

    @patch("documents.views.approve_all_flags_async.delay")
    def test_large_approve_all_is_queued(self, delay_mock):
        job = Job.objects.create(
            owner=self.user,
            filename="bulk.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.NEEDS_REVIEW,
            upload=SimpleUploadedFile("bulk.pdf", b"%PDF-1.4 fake"),
        )
        page = PageResult.objects.create(
            job=job,
            page_number=1,
            extracted_text="Bulk review document",
            tables_json=[],
        )
        ReviewFlag.objects.bulk_create(
            [
                ReviewFlag(
                    page_result=page,
                    word=f"token-{idx}",
                    confidence=0.4,
                    span_offset=idx * 3,
                    span_length=2,
                )
                for idx in range(ASYNC_APPROVE_ALL_THRESHOLD)
            ]
        )

        response = self.client.post(
            reverse("approve_flags", args=[job.id]),
            {"approve_all": "1"},
        )

        self.assertEqual(response.status_code, 302)
        delay_mock.assert_called_once()
        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.REVIEWING)
        self.assertEqual(ReviewFlag.objects.filter(page_result__job=job, reviewed=False).count(), ASYNC_APPROVE_ALL_THRESHOLD)

    @patch("documents.job_control.current_app.control.revoke")
    def test_stop_processing_moves_running_job_to_stopping(self, revoke_mock):
        job = Job.objects.create(
            owner=self.user,
            filename="running.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PROCESSING,
            processing_task_id="celery-task-1",
            upload=SimpleUploadedFile("running.pdf", b"%PDF-1.4 fake"),
        )

        response = self.client.post(reverse("stop_job", args=[job.id]))

        self.assertEqual(response.status_code, 302)
        revoke_mock.assert_called_once_with("celery-task-1", terminate=False)
        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.STOPPING)

    @patch("documents.job_control.current_app.control.revoke")
    def test_stop_processing_moves_pending_job_to_stopped(self, revoke_mock):
        job = Job.objects.create(
            owner=self.user,
            filename="queued.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PENDING,
            processing_task_id="celery-task-2",
            upload=SimpleUploadedFile("queued.pdf", b"%PDF-1.4 fake"),
        )

        response = self.client.post(reverse("stop_job", args=[job.id]))

        self.assertEqual(response.status_code, 302)
        revoke_mock.assert_called_once_with("celery-task-2", terminate=False)
        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.STOPPED)
        self.assertEqual(job.processing_task_id, "")

    @patch("documents.views.queue_processing_job")
    def test_restart_failed_job_requeues_processing(self, queue_mock):
        job = Job.objects.create(
            owner=self.user,
            filename="failed.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.FAILED,
            total_pages=3,
            pages_done=1,
            error_message="Azure timed out",
            upload=SimpleUploadedFile("failed.pdf", b"%PDF-1.4 fake"),
        )

        def fake_queue(job_to_queue):
            job_to_queue.processing_task_id = "new-task-id"
            job_to_queue.save(update_fields=["processing_task_id"])
            return SimpleNamespace(id="new-task-id")

        queue_mock.side_effect = fake_queue

        response = self.client.post(reverse("restart_job", args=[job.id]))

        self.assertEqual(response.status_code, 302)
        queue_mock.assert_called_once()
        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.PENDING)
        self.assertIsNone(job.total_pages)
        self.assertEqual(job.pages_done, 0)
        self.assertEqual(job.error_message, "")
        self.assertEqual(job.processing_task_id, "new-task-id")

    @patch("tasks.ocr_task._run_pipeline")
    def test_process_document_clears_previous_artifacts_before_restart_run(self, run_pipeline_mock):
        job = Job.objects.create(
            owner=self.user,
            filename="restart.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.PENDING,
            processing_task_id="old-task-id",
            upload=SimpleUploadedFile("restart.pdf", b"%PDF-1.4 fake"),
        )
        page = PageResult.objects.create(
            job=job,
            page_number=1,
            extracted_text="Old OCR output",
            tables_json=[],
        )
        ReviewFlag.objects.create(page_result=page, word="OLD", confidence=0.4)

        xlsx_path = os.path.join(self.export_dir, f"{job.id}.xlsx")
        docx_path = os.path.join(self.export_dir, f"{job.id}.docx")
        for path in (xlsx_path, docx_path):
            with open(path, "wb") as handle:
                handle.write(b"stale")

        observed = {}

        def fake_run_pipeline(job_to_run):
            observed["page_count"] = PageResult.objects.filter(job=job_to_run).count()
            observed["flag_count"] = ReviewFlag.objects.filter(page_result__job=job_to_run).count()
            observed["xlsx_exists"] = os.path.exists(xlsx_path)
            observed["docx_exists"] = os.path.exists(docx_path)
            job_to_run.status = Job.Status.COMPLETED
            job_to_run.completed_at = timezone.now()
            job_to_run.export_path = self.export_dir
            job_to_run.processing_task_id = ""
            job_to_run.save(update_fields=["status", "completed_at", "export_path", "processing_task_id"])

        run_pipeline_mock.side_effect = fake_run_pipeline

        process_document.run(str(job.id))

        self.assertEqual(observed["page_count"], 0)
        self.assertEqual(observed["flag_count"], 0)
        self.assertFalse(observed["xlsx_exists"])
        self.assertFalse(observed["docx_exists"])

    def test_process_document_honors_existing_stop_request(self):
        job = Job.objects.create(
            owner=self.user,
            filename="stop-me.pdf",
            document_type=Job.DocumentType.GENERIC,
            status=Job.Status.STOPPING,
            processing_task_id="stop-task-id",
            upload=SimpleUploadedFile("stop-me.pdf", b"%PDF-1.4 fake"),
        )
        page = PageResult.objects.create(
            job=job,
            page_number=1,
            extracted_text="Partial output",
            tables_json=[],
        )
        ReviewFlag.objects.create(page_result=page, word="partial", confidence=0.3)

        process_document.run(str(job.id))

        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.STOPPED)
        self.assertEqual(job.processing_task_id, "")
        self.assertEqual(PageResult.objects.filter(job=job).count(), 0)
        self.assertEqual(ReviewFlag.objects.filter(page_result__job=job).count(), 0)
